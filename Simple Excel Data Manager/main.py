from tkinter import *
import tkinter as tk
from tkinter import messagebox
import openpyxl
import pathlib

file_path = "data.xlsx"
file = pathlib.Path(file_path)
if not file.exists():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["Name", "Surname", "Age", "Job"])  # Başlıklar
    wb.save(file_path)

def submit():
    name = nameValue.get()
    surname = SurnameValue.get()
    age = ageValue.get()
    job = JobValue.get()

    if name == "" or surname == "" or age == "" or job == "":
        messagebox.showerror("Error", "All fields are required!")
        return

    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    sheet.append([name, surname, age, job])
    wb.save(file_path)

    messagebox.showinfo("Success", "Data added successfully!")
    clear()

def show_data():
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    data_window = Toplevel(root)
    data_window.title("Saved Data")
    data_window.geometry("400x300")

    text = Text(data_window, wrap=NONE, font="Arial 10")
    text.pack(expand=True, fill=BOTH)

    scrollbar = Scrollbar(data_window, command=text.yview)
    scrollbar.pack(side=RIGHT, fill=Y)
    text.config(yscrollcommand=scrollbar.set)

    text.insert(END, "Name\tSurname\tAge\tJob\n")
    text.insert(END, "-" * 40 + "\n")

    for row in sheet.iter_rows(values_only=True):
        text.insert(END, f"{row[0]}\t{row[1]}\t{row[2]}\t{row[3]}\n")

# ❌ Silme fonksiyonu (Girilen isim ve soyisme göre)
def delete_entry():
    name = nameValue.get()
    surname = SurnameValue.get()

    if name == "" or surname == "":
        messagebox.showerror("Error", "Enter Name and Surname to delete!")
        return

    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    
    new_rows = [rows[0]]  # Başlıkları sakla

    deleted = False
    for row in rows[1:]:
        if row[0] == name and row[1] == surname:
            deleted = True
        else:
            new_rows.append(row)

    if deleted:
        wb.remove(wb.active)  # Eski sayfayı sil
        new_sheet = wb.create_sheet(title="Sheet1", index=0)
        for row in new_rows:
            new_sheet.append(row)
        wb.save(file_path)
        messagebox.showinfo("Success", "Data deleted successfully!")
        clear()
    else:
        messagebox.showerror("Error", "No matching record found!")

# 🔄 Clear fonksiyonu (Formu temizleme)
def clear():
    nameValue.set("")
    SurnameValue.set("")
    ageValue.set("")
    JobValue.set("")

# 🎨 Arayüz tasarımı
root = Tk()
root.geometry("800x550+250+100")  # Daha geniş pencere
root.title("Data Entry Form")
root.resizable(False, False)
root.configure(bg="#1e3d59")  # Modern mavi tonları

Label(root, text="Please fill out this entry form:", font="Arial 16 bold", bg="#1e3d59", fg="#f5f5f5").place(x=250, y=30)

# 🏷️ Form Etiketleri
Label(root, text="Name", font="Arial 12", bg="#1e3d59", fg="#f5f5f5").place(x=180, y=100)
Label(root, text="Surname", font="Arial 12", bg="#1e3d59", fg="#f5f5f5").place(x=180, y=150)
Label(root, text="Age", font="Arial 12", bg="#1e3d59", fg="#f5f5f5").place(x=180, y=200)
Label(root, text="Job", font="Arial 12", bg="#1e3d59", fg="#f5f5f5").place(x=180, y=250)

# 📝 Form Giriş Alanları
nameValue = StringVar()
SurnameValue = StringVar()
ageValue = StringVar()
JobValue = StringVar()

entry_style = {"width": 40, "bd": 3, "font": "Arial 12"}

nameentry = Entry(root, textvariable=nameValue, **entry_style)
surnameentry = Entry(root, textvariable=SurnameValue, **entry_style)
ageentry = Entry(root, textvariable=ageValue, **entry_style)
jobentry = Entry(root, textvariable=JobValue, **entry_style)

nameentry.place(x=300, y=100)
surnameentry.place(x=300, y=150)
ageentry.place(x=300, y=200)
jobentry.place(x=300, y=250)

# 🎛️ Butonlar
btn_style = {"bg": "#ff6b6b", "fg": "#fff", "font": "Arial 12 bold", "width": 15, "height": 2}

Button(root, text="Submit", command=submit, **btn_style).place(x=100, y=400)
Button(root, text="Clear", command=clear, **btn_style).place(x=250, y=400)
Button(root, text="Delete", command=delete_entry, **btn_style).place(x=400, y=400)
Button(root, text="Show Data", command=show_data, **btn_style).place(x=550, y=400)
Button(root, text="Exit", command=root.destroy, **btn_style).place(x=300, y=470)

root.mainloop()
